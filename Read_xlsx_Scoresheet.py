__author__ = 'joel'
def importfromxls(allxlsxfilepaths):
    """get all the files under a directory and its subdirectories"""
    import os
    import openpyxl
    imported_xls_data = {}

    ###############################################################################################
    ###############################################################################################
    #################################initalize global varibles#############################
    ###############################################################################################
    ###############################################################################################

    #a database of all the dictionary entries
    index = []


    ###############################################################################################
    ###############################################################################################
    #################################Get data into a dictionary array#############################
    ###############################################################################################
    ###############################################################################################

    for filepointer in range(len(allxlsxfilepaths)):
        from openpyxl import load_workbook
        workbook = load_workbook(allxlsxfilepaths[filepointer])
        sheet = workbook['Data Sheet']

        #temporarly store data from spreadsheet
        temp = []
        flag = 1

        #loop for every row in the spreadsheet
        for row in sheet.iter_rows(row_offset=2):
            #loop for every cell in each row
            for cell in row:
                if cell.value == "Subject#.Video.Filename":#break loop once you hit the end of data entry within the spreadsheet.
                    break

                if cell.value == "|":#break loop if you are at the end of the list
                    break

                if cell.value <> None:#get rid of the empty cells
                    temp2 = str(cell.value)#convert to string
                    if temp2[-1] <> ':':#test to see if it is a label, if so don't record
                        temp.append(cell.value)

            if cell.value == "Subject#.Video.Filename":#break loop if you are at the end of the list
                break

            if flag >= 6:
                imported_xls_data[temp[0]]=temp[1:]
                index.append(temp[0])
                temp = []
                flag = 0
            flag= flag + 1


    ###############################################################################################
    ###############################################################################################
    ###############################################################################################
    ###############################################################################################
    ###############################################################################################
    return imported_xls_data, index