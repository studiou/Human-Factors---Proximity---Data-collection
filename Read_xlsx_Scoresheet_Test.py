__author__ = 'joel'

"""get all the files under a directory and its subdirectories"""
import os
import openpyxl
import get_all_xlsx_filepaths

imported_xls_data = {}
filepath = '/Users/joel/Misc/JoelsStuff/OneDrive - GE Appliances/GEA Usability Team Documents/Exploratory Research/Proximity/Prox Python Code/testie'

print str(len(filepath))
#search filepath directory for xlsx files. returns the complete file paths for xlsx files
allxlsxfilepaths = get_all_xlsx_filepaths.builddirectorylist(filepath)


#This all the data from the excel spreadsheet. ie. raw data
imported_xls_data = {}
cleaned_allxlsxfilepaths = []

#get data from excel spreadsheet and dump it into imported_xls_dta array. Just pass the file path to the function and it returns an array with all data from the array
imported_xls_data = allxlsxfilepaths


cleaned_allxlsxfilepaths = []
for item in imported_xls_data:
    if item[len(filepath)+1:len(filepath)+2] != "~":
        if item[-5:] == ".xlsx":
            cleaned_allxlsxfilepaths.append(item)

allxlsxfilepaths = cleaned_allxlsxfilepaths
print allxlsxfilepaths

#allxlsxfilepaths = '/Users/joel/Misc/JoelsStuff/OneDrive - GE Appliances/GEA Usability Team Documents/Exploratory Research/Proximity/Prox Python Code//P04.proximity.score.sheet.xlsx'
###############################################################################################
###############################################################################################
#################################initalize global varibles#############################
###############################################################################################
###############################################################################################

#a database of all the dictionary entries
index = []

###############################################################################################
###############################################################################################
#################################Get data into a dictionary array#############################
###############################################################################################
###############################################################################################

for filepointer in range(len(allxlsxfilepaths)):
    from openpyxl import load_workbook
    workbook = load_workbook(allxlsxfilepaths[filepointer])
    sheet = workbook['Data Sheet']

    print "currently working on sheet " + str(workbook)

    #temporarly store data from spreadsheet
    temp = []
    flag = 1

    #loop for every row in the spreadsheet
    for row in sheet.iter_rows(row_offset=2):
        #loop for every cell in each row
        print "Currently reading row " + str(row) + "currently working on sheet " + str(workbook)

        for cell in row:
            if cell.value == "Subject#.Video.Filename":#break loop once you hit the end of data entry within the spreadsheet.
                break

            if cell.value == "|":#break loop if you are at the end of the list
                break

            if cell.value <> None:#get rid of the empty cells
                temp2 = str(cell.value)#convert to string
                if temp2[-1] <> ':':#test to see if it is a label, if so don't record
                    temp.append(cell.value)

        if cell.value == "Subject#.Video.Filename":#break loop if you are at the end of the list
            break

        if flag >= 6:

            #same video name test
            rename = 0
            for item in index:

                if item == temp [0]:
                    temp[0] = temp[0] + "." + str(rename + 1)
                    rename = rename + 1


            imported_xls_data[temp[0]]=temp[1:]
            index.append(temp[0])
            temp = []
            flag = 0
        flag= flag + 1


        ###############################################################################################
        ###############################################################################################
        ###############################################################################################
        ###############################################################################################
        ###############################################################################################
